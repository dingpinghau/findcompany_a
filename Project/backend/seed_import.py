"""One-time import of legacy case data from 標案案件追縱_20260806 (1).xlsx
into the app database. Skips template/placeholder rows and rows already
imported (matched by project name), so it is safe to re-run.

Usage: python -m backend.seed_import
"""

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from sqlmodel import Session, select

from backend.database import engine, init_db
from backend.models import STAGE_DEFINITIONS, STATUS_OPTIONS, Project, ProjectStage

XLSX_PATH = Path(__file__).resolve().parent.parent / "標案案件追縱_20260806 (1).xlsx"
SHEET_NAME = "標案案件追蹤"
FIRST_DATA_ROW = 31
LAST_DATA_ROW = 54
TEMPLATE_ITEM_NUMBERS = {"範例1", "範例2"}
BLANK_TOKENS = {"n/a", "免", "x", ""}

ITEM_NO_COL = 1
NAME_COL = 3
BUSINESS_UNIT_COL = 4
SALES_REP_COL = 5
NO_GO_REASON_COL = 6
PROGRESS_NOTES_COL = 7
BUDGET_COL = 8
ESTIMATED_BID_COL = 9
ESTIMATED_COST_COL = 10
STATUS_COL = 12
BID_DATE_COL = 34

# (stage_key, planned_date column, actual_date column, overdue_reason column)
STAGE_COLUMNS = [
    ("collect_docs", 13, 14, 15),
    ("kickoff_meeting", 16, 17, 18),
    ("solution_confirm", 19, 20, 21),
    ("benefit_calc", 22, 23, 24),
    ("review_meeting", 25, 26, 27),
    ("sales_signoff", 28, 29, 30),
    ("approval_meeting", 31, 32, 33),
]


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).replace("\xa0", "").strip()
    if text.lower() in BLANK_TOKENS:
        return None
    # Legacy rows store dates as ROC-calendar text, e.g. "115/3/4" -> 2026-03-04
    match = re.match(r"^(\d{2,3})[/.](\d{1,2})[/.](\d{1,2})$", text)
    if not match:
        return None
    roc_year, month, day = (int(g) for g in match.groups())
    try:
        return date(roc_year + 1911, month, day)
    except ValueError:
        return None


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def map_status(raw_status, today, bid_date):
    """Returns (status, assumption_note). assumption_note is None when the
    sheet's status maps cleanly onto STATUS_OPTIONS."""
    if raw_status:
        status = str(raw_status).strip()
        if status in STATUS_OPTIONS:
            return status, None
        return "進行中", f"[原狀態：{status}]"
    if bid_date and bid_date < today:
        return "已結案", "[匯入自舊資料，無明確狀態紀錄，依投標日已過推定為已結案]"
    return "待公告", "[匯入自舊資料，無明確狀態紀錄]"


def run_import():
    init_db()
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    today = date.today()
    stage_defs = {d["stage_key"]: d for d in STAGE_DEFINITIONS}
    bid_definition = stage_defs["bid_submit"]

    imported = 0
    skipped_existing = 0
    with Session(engine) as session:
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            item_no = ws.cell(row=row, column=ITEM_NO_COL).value
            name = clean_text(ws.cell(row=row, column=NAME_COL).value)
            if not name or str(item_no) in TEMPLATE_ITEM_NUMBERS:
                continue

            existing = session.exec(select(Project).where(Project.name == name)).first()
            if existing:
                skipped_existing += 1
                continue

            bid_date = parse_date(ws.cell(row=row, column=BID_DATE_COL).value)
            raw_status = ws.cell(row=row, column=STATUS_COL).value
            status, assumption_note = map_status(raw_status, today, bid_date)

            progress_notes = clean_text(ws.cell(row=row, column=PROGRESS_NOTES_COL).value)
            if assumption_note:
                progress_notes = f"{assumption_note} {progress_notes or ''}".strip()

            project = Project(
                name=name,
                business_unit=clean_text(ws.cell(row=row, column=BUSINESS_UNIT_COL).value),
                sales_rep=clean_text(ws.cell(row=row, column=SALES_REP_COL).value),
                status=status,
                budget_amount=ws.cell(row=row, column=BUDGET_COL).value,
                estimated_bid_amount=ws.cell(row=row, column=ESTIMATED_BID_COL).value,
                estimated_cost=ws.cell(row=row, column=ESTIMATED_COST_COL).value,
                no_go_reason=clean_text(ws.cell(row=row, column=NO_GO_REASON_COL).value),
                progress_notes=progress_notes,
            )
            session.add(project)
            session.commit()
            session.refresh(project)

            for stage_key, planned_col, actual_col, reason_col in STAGE_COLUMNS:
                definition = stage_defs[stage_key]
                session.add(
                    ProjectStage(
                        project_id=project.id,
                        stage_key=stage_key,
                        stage_name=definition["stage_name"],
                        sequence=definition["sequence"],
                        offset_days=definition["offset_days"],
                        planned_date=parse_date(ws.cell(row=row, column=planned_col).value),
                        actual_date=parse_date(ws.cell(row=row, column=actual_col).value),
                        overdue_reason=clean_text(ws.cell(row=row, column=reason_col).value),
                    )
                )
            session.add(
                ProjectStage(
                    project_id=project.id,
                    stage_key="bid_submit",
                    stage_name=bid_definition["stage_name"],
                    sequence=bid_definition["sequence"],
                    offset_days=bid_definition["offset_days"],
                    planned_date=None,
                    actual_date=bid_date,
                    overdue_reason=None,
                )
            )
            session.commit()
            imported += 1

    print(f"Imported {imported} project(s), skipped {skipped_existing} already-imported project(s).")


if __name__ == "__main__":
    run_import()
